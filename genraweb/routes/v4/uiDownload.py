"""Download data"""

import csv
import json
import time
import urllib
from collections import defaultdict, namedtuple
from io import StringIO
from pathlib import Path
from tempfile import NamedTemporaryFile

from flask import Response, request
from flask_openapi3 import APIBlueprint
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from genraweb.defs import FILTER
from genraweb.lib.aggregator.aggregators import METADATA, Aggregator
from genraweb.lib.api_spec.load_api_spec import api_spec_path
from genraweb.lib.chem_id import ChemID
from genraweb.lib.download.radialview import nn_radial_image
from genraweb.lib.fp.fpclass import FPGen
from genraweb.lib.fp.fputils import fp_hybrid_name_from_lists, is_hybrid_fp, parse_fp
from genraweb.lib.logging import logger
from genraweb.lib.misc import check_params
from genraweb.lib.mongofp_NN import searchFP
from genraweb.lib.properties.physprop import PHYSPROP, chem_props
from genraweb.lib.state import GenRAState
from genraweb.resources import V4_URL_PREFIX
from genraweb.routes.api_models import DownloadBody, DownloadPath
from genraweb.routes.api_tags import uiv4_tag

uiDownload_bp = APIBlueprint("uiDownload_bp", __name__)

Field = namedtuple("Field", "name source")
ValDesc = namedtuple("ValDesc", "value description")

FIELDS = [
    Field("chem_id", "chem_id"),
    Field("role", "role"),
    Field("preferred name", "name"),
    Field("dsstox_sid", "dsstox_sid"),
    Field("dsstox_cid", "dsstox_cid"),
    # Field("molecular weight", "mol_weight"),  # covered by PhysChem
    Field("similarity", "similarity"),
]


def hybrid_description(fp):
    """Return description of fingerprint, including hybrid fingerprints"""
    if is_hybrid_fp(fp):
        return "Custom hybrid fingerprint:\n" + ",\n".join(
            f"{FPGen.FPClass.get(fp_id).name or 'x'} with weight={weight}"
            for fp_id, weight in zip(*parse_fp(fp))
        )
    else:
        return FPGen.FPClass.get(fp).name


def physprop_getter(physprop_data, physprop):
    """Function factory to get row data for a physical property, getter functions to be
    added to the FIELDS list.  Called with chem_id, NN map dataframe, column number
    """

    def getter(chem_id, nndf, col, physprop_data=physprop_data, physprop=physprop):
        return physprop_data.get(chem_id, {}).get(physprop.id)

    return getter


def _hide_column(agg, col_def):
    """Should a column be hidden?"""
    return (
        agg.state.get("chem_inc")
        and {"chem_id": col_def["chem_id"], "isChecked": True} not in agg.state.chem_inc
    )


def get_base_data(data):
    """Returns the base pandas DataFrame table, with the chemicals as column
    headers and the assays (and other interest fields) as rows headers.

    Args:
    ----
        data (dict): the POST request JSON data object

    Returns:
    -------
    pandas.DataFrame: see above
    """
    state = GenRAState(data, defaults={"engine": "genrapy"})
    state.chem_id, _ = ChemID.promote_id(state.chem_id)

    agg = Aggregator.aggregator_for(state.summarise, state.sumrs_by)(state)
    # Get predictions if needed.
    if state.get("rra"):
        agg.do_prediction()

    out_rows = []

    # Add special fields from FIELDS
    for field in FIELDS:
        out_row = [field.name]
        out_rows.append(out_row)
        for col_def in agg.frame.col_def:
            if _hide_column(agg, col_def):
                continue
            exports = col_def.get("_exports", [])
            if field.name == "chem_id":  # All export names on first row
                out_row.extend(i["name"] for i in exports)
            else:
                out_row.append(col_def.get(field.source))  # First export
                out_row.extend([None] * (len(exports) - 1))  # Rest of exports

    # Add PhysProp data
    physprop_data = chem_props(agg.frame.col_attr("chem_id"))
    for prop in PHYSPROP:
        out_row = [f"{prop.name} {prop.units}"]
        out_rows.append(out_row)
        for col_def in agg.frame.col_def:
            if _hide_column(agg, col_def):
                continue
            exports = col_def.get("_exports", [])
            if ChemID.id_type(col_def["chem_id"]) == ChemID.CID:
                property = physprop_data[col_def["chem_id"]].get(prop.id)
            else:
                property = None
            out_row.append(property)
            out_row.extend([None] * (len(exports) - 1))  # Rest of exports

    header_n = len(out_rows)

    # Add rows
    for row_def, row in zip(agg.frame.row_def, agg.frame.row):
        if (
            agg.state.get("filter")
            and agg.state.filter.lower() not in row_def["name"].lower()
        ):
            continue  # Show only filtered rows
        out_row = [row_def["name"]]
        out_rows.append(out_row)
        for col_def, cell in zip(agg.frame.col_def, row):
            if _hide_column(agg, col_def):
                continue
            exports = col_def.get("_exports", [])
            out_row.extend(
                next(  # For this export, the first non-null source value
                    (
                        value
                        for export_source in export["source"]
                        if (value := cell.get(export_source)) is not None
                    ),
                    None,
                )
                for export in exports
            )

    return out_rows, header_n


def floaty(data):
    """See if a string can be a float"""
    try:
        return float(data)
    except (ValueError, TypeError):
        return data


def panel_four_metadata(data: dict, tbl: list[list]) -> dict[str, ValDesc]:
    """Make a field name -> value:description mapping."""
    fp = fp_hybrid_name_from_lists(data)
    val_desc = {
        "run_at": ValDesc(time.asctime(), "UTC (GMT) time data was generated."),
        "target": ValDesc(data.get("chem_id"), "Target chemical for predictions."),
        "predict": ValDesc(
            bool(data.get("rra")),
            "True: 'Run Read-Across' done, False: 'Run Read-Across' not done.",
        ),
        "pos_min": ValDesc(
            data.get("pos0"),
            "Minimum positive effect observations required "
            "for a positive effect prediction.",
        ),
        "neg_min": ValDesc(
            data.get("neg0"),
            "Minimum negative effect observations required"
            "for a negative effect prediction.",
        ),
        "s0": ValDesc(data.get("s0"), "Minimum similarity for neighbor selection."),
        "k0": ValDesc(data.get("k0"), "Maximum neighbors to select."),
        "fp_id": ValDesc(data.get("fp"), "Fingerprint used to select neighbors (id)."),
        "filter_by": ValDesc(
            FILTER[data["sel_by"]]["name"],
            FILTER[data["sel_by"]]["description"],
        ),
        "summarise": ValDesc(data["summarise"], "Data Group"),
        "sumrs_by": ValDesc(data["sumrs_by"], "Report/predict"),
    }
    if data.get("engine"):
        val_desc["engine"] = ValDesc(
            data["engine"], "Prediction engine used, legacy genrapred or genrapy."
        )

    # Add metadata for column headers.
    for col_name in tbl[0]:
        if col_name in METADATA:
            val_desc[col_name] = ValDesc("(in data table)", METADATA[col_name])

    if not is_hybrid_fp(fp):
        val_desc["fp_name"] = ValDesc(
            "User defined"
            if fp in ("multitarget", "user-defined")
            else FPGen.FPClass.get(fp).name,
            "Fingerprint used to select neighbors (name).",
        )
    else:
        val_desc["fp_name"] = ValDesc(
            "Custom hybrid fingerprint:\n"
            + ",\n".join(
                f"{FPGen.FPClass.get(fp_id).name or 'x'} with weight={weight}"
                for fp_id, weight in zip(*parse_fp(fp))
            ),
            "Fingerprint used to select neighbors (name).",
        )
        val_desc["fp_weight"] = ValDesc(
            data.get("fp_weight"),
            "Weights used for base fingerprints in custom hybrid.",
        )

    return val_desc


def add_duplicate_header(tbl: list[list], header_n: int) -> None:
    """Duplicate the header row for the table.

    Excel "Tables" (see "create table" below) can't have duplicate headers, so append
    unique numbers.
    """
    field_n = defaultdict(lambda: 0)
    fields = []
    for field in tbl[0]:
        field_n[field] += 1
        header = field
        if field_n[field] != 1:
            header += str(field_n[field])
        fields.append(header)
    tbl[header_n:header_n] = [fields]


def excel_data(tbl: list[list], header_n: int, data: dict) -> bytes:
    """Generate Excel workbook, return bytes."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "GenRA"

    add_duplicate_header(tbl, header_n)
    header_n += 1

    # create workbook
    for row in tbl:
        ws1.append([floaty(i) for i in row])

    # create styles
    row_name = NamedStyle("row_name")
    row_name.font = Font(bold=True)
    target_data = row_name
    preferred_name = NamedStyle("preferred_name")
    preferred_name.font = Font(bold=True, color="4F81BD")
    no_data = Font(color="AAAAAA")
    bottom = Border(bottom=Side(style="thin", color="000000"))

    # create table
    last_letter = get_column_letter(len(tbl[0]))
    ref = f"A{header_n}:{last_letter}{len(tbl)}"
    logger.info(ref)
    tab = Table(displayName="GenRA", ref=ref)
    ws1.add_table(tab)

    # apply styles
    for row_i in range(header_n):
        ws1.cell(row_i + 1, 0 + 1).style = row_name
    for row_i in range(header_n, len(tbl)):
        ws1.cell(row_i + 1, 1 + 1).style = target_data
    name_row = [i[0] for i in tbl].index("preferred name")
    for col_i in range(1, len(tbl[0])):
        ws1.cell(name_row + 1, col_i + 1).style = preferred_name
    width = defaultdict(lambda: 15)
    width[0] = 25
    # if data.get("rra"):
    #     width[1] = 25
    for col_i in range(len(tbl[0])):
        ws1.cell(header_n - 1 + 1, col_i + 1).border = bottom
        for row_i in range(len(tbl)):
            ws1.column_dimensions[get_column_letter(col_i + 1)].width = width[col_i]
            if ws1.cell(row_i + 1, col_i + 1).value == "no_data":
                ws1.cell(row_i + 1, col_i + 1).font = no_data
    ws1.freeze_panes = ws1.cell(header_n + 1, 1 + 1)

    # add metadata
    ws2 = wb.create_sheet(title="Metadata")
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 20
    fp = fp_hybrid_name_from_lists(data)
    val_desc = panel_four_metadata(data, tbl)
    for key, (val, desc) in val_desc.items():
        ws2.append([key, val, desc])

    if is_hybrid_fp(fp):
        # ws2.append() appends after the last completely untouched row, so if this had
        # been done above, data would be placed after the rows touched here.
        key_i = list(val_desc).index("fp_name")
        ws2.cell(key_i + 1, 1 + 1).alignment = Alignment(wrap_text=True)
        ws2.row_dimensions[key_i + 1].height = 15 * (2 + data.get("fp").count(","))
        ws2.column_dimensions["B"].width = 40

    # return data
    with NamedTemporaryFile() as out:
        wb.save(out.name)
        return Path(out.name).read_bytes()


def add_csv_metadata(tbl: list, data: dict) -> None:
    """Add metadata column to tbl."""
    val_desc = panel_four_metadata(data, tbl)
    metadata = ["metadata"] + [
        f"{k}: {v.value} -- {v.description}" for k, v in val_desc.items()
    ]
    for row in range(len(tbl)):
        tbl[row][:0] = [metadata[row]] if row < len(metadata) else [None]


@uiDownload_bp.post(
    urllib.parse.urljoin(V4_URL_PREFIX, "uiDownload/<ftype>"),
    summary=DownloadBody.__doc__,
    tags=[uiv4_tag],
    responses={200:{"description":"Details of tox for neighbouring chemicals"}}
)
def uiDownload(path: DownloadPath, body: DownloadBody) -> Response:
    """Run GenRA analysis and prediction, download results."""
    data = body.model_dump()
    ftype = path.ftype
    # probably need better dispatch handling here, but ok for now
    if ftype == "allNN":
        return top_100_nn(data)
    elif ftype == "RAview":
        return ra_view(data)

    tbl, header_n = get_base_data(data)

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    if ftype == "csv":
        mimetype = "text/csv"
        add_csv_metadata(tbl, data)
        out = StringIO()
        csv.writer(out).writerows(tbl)
        response_data = out.getvalue()
        filename = f"genra_{timestamp}.csv"
    else:
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response_data = excel_data(tbl, header_n, data)
        filename = f"genra_{timestamp}.xlsx"

    logger.info("Generated %s", filename)
    return Response(
        response_data,
        mimetype=mimetype,
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


def top_100_nn(data):
    """CSV download of top 100 nearest neighbors"""
    check_params(
        data,
        ["chem_id", "fp", "sel_by"],
        optional=["fp_weight"],
    )

    chem_id = data.get("chem_id")
    fp = fp_hybrid_name_from_lists(data)
    sel_by = data.get("sel_by")

    neighbors = searchFP(chem_id, fp=fp, sel_by=sel_by, s0=0, max_hits=100)

    header = {
        "metadata": None,
        "chem_id": "ID of neighbor",
        "similarity": "Similarity of target chemical to neighbor",
        "name": "Name of neighbor",
        "dsstox_sid": "DSSTox Substance ID of neighbor",
        "dsstox_cid": "DSSTox Chemical ID of neighbor",
        "casrn": "CASRN of neighbor",
        "mol_weight": "Molecular weight of neighbor",
        "smiles": "SMILES of neighbor",
    }
    # Include individual similarity for hybrid FP, and non-hybrid, as column header
    # provides a hint as to the FP type.
    # FIXME - select without using "jaccard_", "euclid_"
    extra = [
        i for i in neighbors[0] if i.startswith("jaccard_") or i.startswith("euclid_")
    ]
    header.update(
        {
            # jaccard_chm_mrgn, euclid_bio_pest => chm_mrgn, bio_pest
            i: f"Similarity for {FPGen.FPClass[i.split('_', 1)[1]].name}"
            for i in extra
        }
    )
    # Include FP columns if not hybrid.
    if not is_hybrid_fp(fp):
        header.update(
            {
                "fp_list": "Comma separated list of FP elements.",
                "fp_bitstring": "FP bitstring.",
            }
        )
        col_names = FPGen.FPClass[fp].bit_names()
        fp_id = FPGen.FPClass[fp].fp_id
        header.update({i: None for i in col_names})
        for neighbor in neighbors:
            # use col_names order
            fpds_key = f"fpds_{fp_id}"
            if isinstance(neighbor[fpds_key], list):
                fp_list = [i for i in col_names if i in neighbor[fpds_key]]
                fp_bits = [1 if i in fp_list else 0 for i in col_names]
                fp_string = "".join(map(str, fp_bits))
            else:
                fp_list = [f"{k}:{v}" for k, v in neighbor[fpds_key].items()]
                fp_bits = list(neighbor[fpds_key].values())
                fp_string = ",".join(map(str, fp_bits))
            fp_list = ",".join(fp_list)
            neighbor |= {"fp_list": fp_list, "fp_bitstring": fp_string}
            neighbor |= dict(zip(col_names, fp_bits))

    # Fill in metadata column
    metadata = [
        ("run_at", time.strftime("%Y-%m-%d %H:%M:%S")),
        ("target", chem_id),
        ("fp_id", data["fp"]),
        ("fp_weight", data.get("fp_weight", 1)),
        (
            "metric_name",
            "Hybrid"
            if "," in data["fp"]
            else FPGen.FPClass[data["fp"]].nn_distance.title(),
        ),
        ("filter_by", sel_by),
        ("fp_name", hybrid_description(fp).replace("\n", " ")),
    ]
    metadata.extend(i for i in header.items() if i[1] is not None)
    for head, row in zip(metadata, neighbors):
        row["metadata"] = f"{head[0]}: {head[1]}"

    out = StringIO()
    writer = csv.DictWriter(out, header, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(neighbors)

    mimetype = "text/csv"
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    filename = f"genra_{timestamp}.csv"
    return Response(
        out.getvalue(),
        mimetype=mimetype,
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


def ra_view(data):
    """PNG radial plot download"""
    check_params(
        data,
        ["chem_id", "fp", "sel_by"],
        optional=["fp_weight"],
    )

    image = nn_radial_image(GenRAState(data))

    mimetype = "image/png"
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    filename = f"genra_{timestamp}.png"
    return Response(
        image,
        mimetype=mimetype,
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
